"""Genera el Reporte_MS_MS_TOP_200_Spotify (Market Share YTD) a partir del
histórico acumulado en src/history.py.

Fórmula real (verificada contra PLANTILLA_SEMANAL_MS_TOP200.xlsx, filas
99-112 — ver claude/mapeo_logica_plantillas.md sección 1):

    % Market Share YTD (sello, país) =
        suma streams_top200 del sello, semanas 1..N del año
        -------------------------------------------------------
        suma streams_top200 de los 7 sellos, semanas 1..N del año

Se compara el mismo número de semanas (1..N) entre el año en curso y el
año anterior. NO es un promedio de porcentajes semanales.

La pestaña resumen "% Market Share" se escribe como una cuadrícula de
tablas, una por país (4 por fila de bloques), replicando el formato visual
de PLANTILLA_SEMANAL_MS_TOP200.xlsx -- ver _escribir_resumen_pct.

Las pestañas individuales por país (una por config.PAISES_MS) son ahora una
cuadrícula semanal: columnas = semanas ya guardadas en
history.chart_track_weekly (fecha + n° de semana), filas = las 5 bandas de
config.BANDAS_MARKET_SHARE, cada una con las 7 filas de sello de
config.LABEL_GROUPS_MS -- replica la sub-tabla "Streams (%)" de
PLANTILLA_SEMANAL_MS_TOP200.xlsx (ver _escribir_pagina_pais / Ajuste 7 en
el plan). Las otras dos sub-tablas de la plantilla real por banda ("Tracks"
y "Streams" crudo) se dejaron fuera a pedido del usuario -- solo la de "%".

Ojo con el histórico disponible: esta cuadrícula solo puede tener columnas
para las semanas que ya se guardaron en chart_track_weekly, que empezó a
llenarse recién en el Ajuste 5 (no tiene sembrado retroactivo de años
anteriores, a diferencia de ms_label_weekly). El usuario decidió a
propósito no importar el histórico ya calculado del archivo real (que sí
llega hasta 2021) -- la cuadrícula arranca vacía y se va llenando semana a
semana hacia adelante, igual que Detalle Tracks.
"""
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, history


def calcular_ytd_por_pais(
    country_code: str, anio_actual: int, hasta_semana: int
) -> pd.DataFrame:
    """% Market Share YTD de un país: un DataFrame con una fila por sello
    (en el orden de config.LABEL_GROUPS_MS), comparando anio_actual vs.
    anio_actual - 1, mismas semanas 1..N en ambos.

    N = min(hasta_semana, semanas realmente disponibles de anio_actual - 1).
    Esto evita comparar, por ejemplo, 25 semanas de 2026 contra solo 24 de
    2025 si el histórico del año anterior no llega tan lejos todavía — la
    comparación YTD deja de tener sentido si no son las mismas semanas.
    """
    anio_anterior = anio_actual - 1

    historico_pais_anterior = history.cargar_ms_label_weekly()
    historico_pais_anterior = historico_pais_anterior[
        (historico_pais_anterior["country_code"] == country_code)
        & (historico_pais_anterior["anio"] == anio_anterior)
    ]
    semana_max_anterior = (
        int(historico_pais_anterior["semana"].max())
        if not historico_pais_anterior.empty
        else 0
    )
    hasta_semana_efectiva = min(hasta_semana, semana_max_anterior)

    ytd_actual = history.query_ytd_ms(anio_actual, hasta_semana_efectiva)
    ytd_actual = ytd_actual[ytd_actual["country_code"] == country_code]
    ytd_anterior = history.query_ytd_ms(anio_anterior, hasta_semana_efectiva)
    ytd_anterior = ytd_anterior[ytd_anterior["country_code"] == country_code]

    streams_actual = ytd_actual.groupby("label_group")["streams_top200"].sum()
    streams_anterior = ytd_anterior.groupby("label_group")["streams_top200"].sum()

    total_actual = streams_actual.sum()
    total_anterior = streams_anterior.sum()

    filas = []
    for label in config.LABEL_GROUPS_MS:
        s_actual = float(streams_actual.get(label, 0.0))
        s_anterior = float(streams_anterior.get(label, 0.0))
        pct_actual = s_actual / total_actual if total_actual else 0.0
        pct_anterior = s_anterior / total_anterior if total_anterior else 0.0
        filas.append({
            "label_group": label,
            f"pct_YTD_{anio_actual}": pct_actual,
            f"pct_YTD_{anio_anterior}": pct_anterior,
            "g_l": pct_actual - pct_anterior,
        })
    return pd.DataFrame(filas)


def calcular_streams_pct_grid(country_code: str) -> pd.DataFrame:
    """Tabla larga (tidy) con el % de streams por banda y sello, semana a
    semana, para un país -- igual a lo que muestran las sub-tablas
    "Streams (%) TOP N" de PLANTILLA_SEMANAL_MS_TOP200.xlsx, pero a partir
    del detalle track-por-track de history.chart_track_weekly (no de un
    agregado guardado aparte -- se calcula al vuelo cada vez que se genera
    el reporte).

    Una fila por (anio, semana, chart_date, banda, label_group, pct). Solo
    incluye las semanas que ya están en chart_track_weekly para ese país
    (ver nota del módulo sobre por qué no hay histórico antes del Ajuste 5).
    """
    tracks = history.cargar_chart_track_weekly()
    tracks = tracks[tracks["country_code"] == country_code].copy()
    if tracks.empty:
        return pd.DataFrame(columns=["anio", "semana", "chart_date", "banda", "label_group", "pct"])

    tracks["stream_count"] = tracks["stream_count"].fillna(0.0)

    filas = []
    for (anio, semana, chart_date), grupo_semana in tracks.groupby(["anio", "semana", "chart_date"]):
        for banda in config.BANDAS_MARKET_SHARE:
            grupo_banda = grupo_semana[grupo_semana["position"] <= banda]
            total = grupo_banda["stream_count"].sum()
            streams_por_label = grupo_banda.groupby("label_group")["stream_count"].sum()
            for label in config.LABEL_GROUPS_MS:
                streams_label = float(streams_por_label.get(label, 0.0))
                pct = streams_label / total if total else 0.0
                filas.append({
                    "anio": anio, "semana": semana, "chart_date": chart_date,
                    "banda": banda, "label_group": label, "pct": pct,
                })
    return pd.DataFrame(filas)


def construir_resumen_pct(anio_actual: int, hasta_semana: int) -> pd.DataFrame:
    """Tabla larga (tidy) con el % Market Share YTD de los 17 países, para
    la pestaña resumen "% Market Share". Una fila por país/sello.
    """
    partes = []
    for country_code in config.PAISES_MS:
        tabla_pais = calcular_ytd_por_pais(country_code, anio_actual, hasta_semana)
        tabla_pais.insert(0, "country_code", country_code)
        partes.append(tabla_pais)
    return pd.concat(partes, ignore_index=True)


# Layout de la pestaña resumen "% Market Share", replicando el formato
# visual de PLANTILLA_SEMANAL_MS_TOP200.xlsx (verificado 1:1 contra ese
# archivo real -- fila 2 banner, fila 4 encabezado de bloque, fila 6
# subencabezado, filas 7-13 sellos, luego 1 fila en blanco y se repite):
#   - Banner "TOP 200 WEEKLY MARKET SHARE" arriba de todo (fila 2), y
#     freeze_panes en A4 para que las filas 1-3 queden siempre visibles al
#     desplazarse hacia abajo (pedido explícito del usuario).
#   - Un bloque de 4 columnas por país (nombre, YTD año actual, YTD año
#     anterior, G/L), con el nombre del país en un banner arriba del bloque
#     y 7 filas de sellos debajo (config.ORDEN_LABELS_MS_RESUMEN). 4 bloques
#     de país por fila, con una columna angosta de separación entre cada
#     uno, igual que la plantilla real.
# NOTA: la plantilla original resalta con colores algunas celdas de G/L,
# pero la regla real (revisada en el archivo) no es simple ("mayor que la
# fila de arriba", comparaciones entre celdas específicas) -- no se pudo
# generalizar con certeza a partir de eso, así que se dejó fuera de este
# ajuste a propósito. Avisar si se quiere una regla más simple (ej. verde si
# G/L > 0, rojo si G/L < 0) para agregarla.
_MS_COL_INICIAL = 2  # columna B, igual que la plantilla real (A queda vacía)
_MS_COLS_POR_BLOQUE = 4  # nombre, YTD actual, YTD anterior, G/L
_MS_BLOQUES_POR_FILA = 4
_MS_FILAS_POR_BLOQUE = 11  # banner + blanco + subencabezado + 7 sellos + blanco
_MS_FILA_BANNER_TOP200 = 2
_MS_FILA_PRIMER_BLOQUE = 4


def _escribir_resumen_pct(ws, anio_actual: int, hasta_semana: int) -> None:
    """Escribe la pestaña resumen "% Market Share" como una cuadrícula de
    tablas por país (ver constantes _MS_* arriba) -- se arma directo con
    openpyxl (no con `.to_excel(...)`) por el mismo motivo que en
    chart_semanal.py: acá hace falta control celda por celda (banners
    combinados, freeze_panes) que `.to_excel(...)` no ofrece.
    """
    negrita_banner = Font(bold=True, color=config.COLOR_BANNER_MS_TEXTO)
    relleno_banner = PatternFill(
        start_color=config.COLOR_BANNER_MS_FONDO, end_color=config.COLOR_BANNER_MS_FONDO, fill_type="solid"
    )
    centrado = Alignment(horizontal="center", vertical="center")
    negrita = Font(bold=True)

    ancho_total_columnas = _MS_BLOQUES_POR_FILA * (_MS_COLS_POR_BLOQUE + 1) - 1
    columna_final = _MS_COL_INICIAL + ancho_total_columnas - 1
    celda_top200 = ws.cell(row=_MS_FILA_BANNER_TOP200, column=_MS_COL_INICIAL, value="TOP 200 WEEKLY MARKET SHARE")
    ws.merge_cells(
        start_row=_MS_FILA_BANNER_TOP200, start_column=_MS_COL_INICIAL,
        end_row=_MS_FILA_BANNER_TOP200, end_column=columna_final,
    )
    celda_top200.font = Font(bold=True, size=14, color=config.COLOR_BANNER_MS_TEXTO)
    celda_top200.fill = relleno_banner
    celda_top200.alignment = centrado

    etiqueta_actual = f"YTD {str(anio_actual)[-2:]}"
    etiqueta_anterior = f"YTD {str(anio_actual - 1)[-2:]}"

    for i, country_code in enumerate(config.ORDEN_PAISES_MS_RESUMEN):
        fila_bloque = i // _MS_BLOQUES_POR_FILA
        col_bloque = i % _MS_BLOQUES_POR_FILA
        fila_header = _MS_FILA_PRIMER_BLOQUE + fila_bloque * _MS_FILAS_POR_BLOQUE
        fila_subheader = fila_header + 2
        fila_primer_dato = fila_subheader + 1
        col_inicio = _MS_COL_INICIAL + col_bloque * (_MS_COLS_POR_BLOQUE + 1)
        col_fin = col_inicio + _MS_COLS_POR_BLOQUE - 1

        nombre_visible = config.NOMBRE_PAIS_MS_RESUMEN[country_code]

        celda_pais = ws.cell(row=fila_header, column=col_inicio, value=nombre_visible)
        ws.merge_cells(start_row=fila_header, start_column=col_inicio, end_row=fila_header, end_column=col_fin)
        celda_pais.font = negrita_banner
        celda_pais.fill = relleno_banner
        celda_pais.alignment = centrado

        encabezados = [nombre_visible, etiqueta_actual, etiqueta_anterior, "G/L"]
        for j, texto in enumerate(encabezados):
            celda = ws.cell(row=fila_subheader, column=col_inicio + j, value=texto)
            celda.font = negrita
            celda.alignment = centrado

        tabla_pais = calcular_ytd_por_pais(country_code, anio_actual, hasta_semana)
        tabla_pais = tabla_pais.set_index("label_group")
        for k, label in enumerate(config.ORDEN_LABELS_MS_RESUMEN):
            r = fila_primer_dato + k
            fila_label = tabla_pais.loc[label] if label in tabla_pais.index else None
            ws.cell(row=r, column=col_inicio, value=label)
            for col_offset, campo in enumerate(
                [f"pct_YTD_{anio_actual}", f"pct_YTD_{anio_actual - 1}", "g_l"], start=1
            ):
                valor = float(fila_label[campo]) if fila_label is not None else None
                celda_valor = ws.cell(row=r, column=col_inicio + col_offset, value=valor)
                celda_valor.number_format = "0.0%"

        for col in range(col_inicio, col_fin + 1):
            ws.column_dimensions[get_column_letter(col)].width = 11
        ws.column_dimensions[get_column_letter(col_fin + 1)].width = 2  # separadora

    ws.freeze_panes = "A4"


# Layout de la pestaña individual de cada país (ver calcular_streams_pct_grid
# y el docstring del módulo). Verificado contra la pestaña "CO" real de
# PLANTILLA_SEMANAL_MS_TOP200.xlsx: columna A = banda (combinada verticalmente
# sobre sus 7 filas de sello), columna B = sello, columnas C+ = una por
# semana. Fila 2 = fecha, fila 3 = n° de semana -- esas dos filas y las
# columnas A/B quedan fijas (freeze_panes), igual que las columnas de tops y
# disquera quedan fijas al desplazarse a la derecha (pedido explícito del
# usuario).
_MSPAIS_COL_BANDA = 1
_MSPAIS_COL_LABEL = 2
_MSPAIS_COL_PRIMERA_SEMANA = 3
_MSPAIS_FILA_TITULO = 1
_MSPAIS_FILA_FECHA = 2
_MSPAIS_FILA_SEMANA = 3
_MSPAIS_FILA_PRIMER_DATO = 4


def _escribir_pagina_pais(ws, country_code: str) -> None:
    """Escribe la pestaña individual de un país como cuadrícula semanal
    (ver constantes _MSPAIS_* arriba). Si todavía no hay ninguna semana
    guardada en chart_track_weekly para este país, la cuadrícula sale sin
    columnas de datos (solo encabezados) -- se va llenando sola a medida
    que se cargan bases nuevas.
    """
    negrita = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    nombre_visible = config.NOMBRE_PAIS_MS_RESUMEN.get(country_code, country_code)
    celda_titulo = ws.cell(
        row=_MSPAIS_FILA_TITULO, column=_MSPAIS_COL_BANDA,
        value=f"{nombre_visible} - % Market Share por banda (Streams)",
    )
    celda_titulo.font = Font(bold=True, size=12)

    grid = calcular_streams_pct_grid(country_code)
    semanas = (
        grid[["anio", "semana", "chart_date"]]
        .drop_duplicates()
        .sort_values(["anio", "semana"])
        .reset_index(drop=True)
    )
    pct_por_celda = {
        (r.anio, r.semana, r.banda, r.label_group): r.pct for r in grid.itertuples(index=False)
    }

    ws.cell(row=_MSPAIS_FILA_SEMANA, column=_MSPAIS_COL_LABEL, value="Etiquetas\nde fila").font = negrita

    for j, s in enumerate(semanas.itertuples(index=False)):
        col = _MSPAIS_COL_PRIMERA_SEMANA + j
        celda_fecha = ws.cell(row=_MSPAIS_FILA_FECHA, column=col, value=pd.Timestamp(s.chart_date))
        celda_fecha.number_format = "yyyy-mm-dd"
        celda_fecha.font = negrita
        celda_fecha.alignment = centrado
        celda_semana = ws.cell(row=_MSPAIS_FILA_SEMANA, column=col, value=int(s.semana))
        celda_semana.font = negrita
        celda_semana.alignment = centrado
        ws.column_dimensions[get_column_letter(col)].width = 11

    fila = _MSPAIS_FILA_PRIMER_DATO
    for banda in config.BANDAS_MARKET_SHARE:
        fila_inicio_banda = fila
        for label in config.LABEL_GROUPS_MS:
            ws.cell(row=fila, column=_MSPAIS_COL_LABEL, value=label)
            for j, s in enumerate(semanas.itertuples(index=False)):
                col = _MSPAIS_COL_PRIMERA_SEMANA + j
                valor = pct_por_celda.get((s.anio, s.semana, banda, label))
                celda_valor = ws.cell(row=fila, column=col, value=valor)
                celda_valor.number_format = "0.0%"
            fila += 1

        celda_banda = ws.cell(row=fila_inicio_banda, column=_MSPAIS_COL_BANDA, value=f"Streams\n(%)\nTOP {banda}")
        ws.merge_cells(
            start_row=fila_inicio_banda, start_column=_MSPAIS_COL_BANDA,
            end_row=fila - 1, end_column=_MSPAIS_COL_BANDA,
        )
        celda_banda.font = negrita
        celda_banda.alignment = centrado
        fila += 1  # fila en blanco separadora entre bandas, igual que la plantilla real

    ws.column_dimensions[get_column_letter(_MSPAIS_COL_BANDA)].width = 12
    ws.column_dimensions[get_column_letter(_MSPAIS_COL_LABEL)].width = 14

    ws.freeze_panes = ws.cell(row=_MSPAIS_FILA_PRIMER_DATO, column=_MSPAIS_COL_PRIMERA_SEMANA).coordinate


def generar_reporte(
    df_semana: pd.DataFrame, output_path: Path, guardar_en_historico: bool = True
) -> Path:
    """Genera el reporte de Market Share a partir del DataFrame de la semana
    nueva (el que devuelve load_data.load_source). Guarda esa semana en el
    histórico (a menos que ya se haya guardado antes, con
    guardar_en_historico=False) y calcula el YTD con todo el histórico
    acumulado hasta esa semana.
    """
    if guardar_en_historico:
        history.append_semana_ms(df_semana)

    fecha = pd.Timestamp(df_semana["chart_date"].unique()[0])
    anio_actual = fecha.year

    todo_el_historico = history.cargar_ms_label_weekly()
    hasta_semana = int(todo_el_historico.loc[todo_el_historico["anio"] == anio_actual, "semana"].max())

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        ws_resumen = writer.book.create_sheet(config.MS_SHEET_PORCENTAJE)
        _escribir_resumen_pct(ws_resumen, anio_actual, hasta_semana)
        for country_code in config.PAISES_MS:
            ws_pais = writer.book.create_sheet(country_code)
            _escribir_pagina_pais(ws_pais, country_code)

    return output_path
