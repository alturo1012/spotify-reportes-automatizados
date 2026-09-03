"""Genera el Reporte_Chart_Top_Semanal_Spotify_Latam a partir del histórico
acumulado en src/history.py (pestaña "Resumen Total") y de la semana nueva
cargada (pestaña "Detalle Tracks").

Ver claude/mapeo_logica_plantillas.md sección 2 para el detalle completo.

ADVERTENCIA (sección 2.2 del mapeo): el conteo "Universal" que usa este
módulo se basa en `label_group == 'Universal'`. En el proceso manual
original, ese conteo salía de contar celdas coloreadas a mano semana a
semana (macro VBA `CountCcolor`), no de una fórmula sobre datos. Es la mejor
aproximación automatizable, pero no hay garantía matemática de que coincida
100% con el criterio manual — validar contra la primera semana real de uso
antes de confiar en el número (Paso 6 del plan).

"Detalle Tracks" y el listado de canciones de "Resumen Total" siguen
mostrando SOLO la semana que se acaba de cargar, no una serie histórica
(decisión confirmada con el usuario) -- pero desde este cambio, el detalle
track por track de cada semana sí se GUARDA aparte en
`history.chart_track_weekly` (vía `history.append_semana_tracks`), para no
perderlo cuando se cargue la semana siguiente. Es guardado "hacia adelante"
nada más: no tiene sembrado retroactivo de semanas anteriores a este cambio
(ver la nota sobre "backfill" en claude/plan_fusion_paso_a_paso.md).
"""
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, history, spotify_release_dates

_RELLENO_ROJO = PatternFill(
    start_color=config.COLOR_SEMAFORO_ROJO, end_color=config.COLOR_SEMAFORO_ROJO, fill_type="solid"
)
_RELLENO_AMARILLO = PatternFill(
    start_color=config.COLOR_SEMAFORO_AMARILLO, end_color=config.COLOR_SEMAFORO_AMARILLO, fill_type="solid"
)
_RELLENO_VERDE = PatternFill(
    start_color=config.COLOR_SEMAFORO_VERDE, end_color=config.COLOR_SEMAFORO_VERDE, fill_type="solid"
)

# Layout de la pestaña "Resumen Total", replicando el formato visual de la
# plantilla original (PLANTILLA_SEMANAL_ChartTop.xlsm / Reporte_Chart_Top
# Semanal Spotify Latam a Sem 24 de 2026.xlsm), verificado columna por
# columna contra ese archivo real:
#   - Columnas A-D fijas (año / mes / semana / separador) + filas 1-6 fijas,
#     con "congelar paneles" en E7 (freeze_panes) para que se vean siempre
#     al desplazarse hacia la derecha o hacia abajo.
#   - Cada país ocupa un bloque de 5 columnas (top10/30/50/100/200), con el
#     nombre del país en una celda combinada en la fila 5 y los números de
#     banda en la fila 6.
#   - Una columna angosta de separación entre cada bloque de país (y entre
#     las columnas fijas y el primer país).
#   - Debajo de la serie histórica (con un salto de filas en blanco) va el
#     "listado de canciones" de la semana actual: mismos bloques de país,
#     pero en vez de un conteo, la posición real de cada canción en el país
#     donde aparece (ver _escribir_listado_canciones).
# NOTA: la plantilla original también trae una fila 1-2 con una leyenda
# "TOP 10/30/50/100/200" y unos valores de referencia (3/9/15/30/60) que en
# su momento no se pudieron confirmar de dónde salían -- ya se explicaron:
# son el objetivo de participación de Universal (30% de cada banda, ver
# config.PCT_OBJETIVO_UNIVERSAL y _color_semaforo) redondeado a entero
# (10/30/50/100/200 x 30% = 3/9/15/30/60 exacto). Esa leyenda en sí (las dos
# filas fijas arriba de todo) sigue sin replicarse -- se puede agregar si
# hace falta, ahora que se sabe qué representa.

_COL_ANIO = 1
_COL_MES = 2
_COL_SEMANA = 3
_COL_SEPARADOR_INICIAL = 4
_COL_PRIMER_PAIS = 5
_FILA_HEADER_PAIS = 5
_FILA_HEADER_BANDA = 6
_FILA_PRIMER_DATO = 7

# Filas en blanco entre la última fila de la serie histórica y el título
# "Week Ending - ..." del listado de canciones (equivale al gran salto de
# filas -- "después de la fila 300" -- que tiene la plantilla original;
# acá no hace falta reservar cientos de filas porque la tabla histórica
# crece semana a semana, así que basta un salto corto y fijo).
_FILAS_ANTES_DE_LISTADO = 2


def construir_resumen_total() -> pd.DataFrame:
    """Serie histórica completa (todo lo acumulado en chart_band_weekly):
    una fila por semana, con columnas por país x banda, en el mismo orden
    que la plantilla original (config.ORDEN_PAISES_CHART x
    config.BANDAS_CHART -- OJO, es un orden de país distinto al de
    Market Share / config.PAISES_MS, ver la constante).
    """
    df = history.cargar_chart_band_weekly()
    if df.empty:
        return pd.DataFrame(columns=["anio", "semana", "mes"])

    tabla = df.pivot_table(
        index=["anio", "semana", "mes"],
        columns=["country_code", "banda"],
        values="conteo_universal",
        aggfunc="first",
    )

    columnas_ordenadas = [
        (pais, banda)
        for pais in config.ORDEN_PAISES_CHART
        for banda in config.BANDAS_CHART
        if (pais, banda) in tabla.columns
    ]
    tabla = tabla[columnas_ordenadas]
    tabla.columns = [f"{pais}_top{banda}" for pais, banda in tabla.columns]
    tabla = tabla.reset_index().sort_values(["anio", "semana"])
    return tabla


def _tier_de_posicion(posicion: int):
    """A qué banda (10/30/50/100/200) pertenece una posición: la más chica
    de config.BANDAS_CHART que la contiene (posición 9 -> banda 10,
    posición 25 -> banda 30, etc.). None si no entra en ninguna (fuera del
    Top 200).
    """
    for banda in config.BANDAS_CHART:
        if posicion <= banda:
            return banda
    return None


def construir_listado_canciones(df_semana: pd.DataFrame) -> pd.DataFrame:
    """Listado de canciones de la semana que se acaba de cargar (NO es
    histórico -- cambia por completo cada vez que se sube una fuente nueva,
    igual que "Detalle Tracks"), una fila por canción con:

    - "cancion": "Título / Artista".
    - "region": tal cual viene de la fuente (Anglo/Latin/...). Se asume que
      es la misma para todos los países donde aparece esa canción -- se usa
      el primer valor visto por canción.
    - Una columna por país donde esa canción aparece, con el número de
      posición real puesto bajo la banda (10/30/50/100/200) a la que
      pertenece esa posición (una sola celda no vacía por país, no las 5).
    - "isrc": ISRC de esa canción (primer valor visto, igual criterio que
      "region"). Uso interno -- no se escribe tal cual en el Excel, sirve
      para resolver "fecha_lanzamiento" (ver agregar_fecha_lanzamiento).
    - "paises_presente": en cuántos países aparece.
    - "suma_posiciones": suma de sus posiciones en todos esos países.

    Orden de filas: por cantidad de países (de mayor a menor) y, para
    empatar, por la suma de posiciones (de menor a mayor) -- así las
    canciones que están en más países y mejor posicionadas quedan primero
    ("las mejores canciones"). Es una decisión razonable, no algo pedido
    explícito -- fácil de cambiar si no es el orden que se espera.

    Solo devuelve las primeras config.TOP_N_LISTADO_CANCIONES (200 por
    defecto) de ese orden -- pedido explícito del usuario, para no listar
    las 1000+ canciones que puede traer una semana completa.
    """
    columnas_vacio = ["cancion", "region", "isrc", "paises_presente", "suma_posiciones"]
    if df_semana.empty:
        return pd.DataFrame(columns=columnas_vacio)

    df = df_semana.copy()
    if "region" not in df.columns:
        # Defensivo: "region"/"ISRC" vienen de la fuente BQ real
        # (config.SOURCE_COLUMNS), pero no todo caller de prueba las incluye
        # -- que falten no debe tumbar el reporte, solo quedan en blanco.
        df["region"] = None
    if "ISRC" not in df.columns:
        df["ISRC"] = None
    df["cancion"] = (
        df["song_name"].astype(str).str.strip() + " / " + df["artist"].astype(str).str.strip()
    )
    df["banda"] = df["position"].apply(_tier_de_posicion)
    df = df.dropna(subset=["banda"])
    if df.empty:
        return pd.DataFrame(columns=columnas_vacio)
    df["banda"] = df["banda"].astype(int)

    resumen_cancion = df.groupby("cancion").agg(
        region=("region", "first"),
        isrc=("ISRC", "first"),
        paises_presente=("country_code", "nunique"),
        suma_posiciones=("position", "sum"),
    )

    posiciones_por_pais = df.pivot_table(
        index="cancion", columns=["country_code", "banda"], values="position", aggfunc="first"
    )
    posiciones_por_pais.columns = [
        f"{pais}_top{banda}" for pais, banda in posiciones_por_pais.columns
    ]

    listado = resumen_cancion.join(posiciones_por_pais).reset_index()
    listado = listado.sort_values(
        ["paises_presente", "suma_posiciones"], ascending=[False, True]
    ).reset_index(drop=True)
    # Solo las mejores config.TOP_N_LISTADO_CANCIONES (por defecto 200) --
    # pedido explícito del usuario, para no listar las 1000+ canciones de
    # una semana completa.
    return listado.head(config.TOP_N_LISTADO_CANCIONES).reset_index(drop=True)


# Motivo por el que la última corrida no pudo resolver fechas de
# lanzamiento (None si salió bien o si todavía no se intentó). Existe porque
# el .exe se empaqueta con --windowed, es decir SIN consola: el `print(...)`
# del aviso no lo ve nadie cuando se usa la app empaquetada. La GUI lee esto
# después de generar para poder mostrarlo en el mensaje final (ver
# gui.generar / App._exito). Bug real: el usuario generó un reporte con el
# .exe, la columna salió vacía, y no tenía forma de saber por qué.
_ULTIMO_AVISO_FECHAS = None


def ultimo_aviso_fechas():
    """Devuelve el motivo del último fallo al resolver fechas de lanzamiento
    (o None si la última corrida salió bien)."""
    return _ULTIMO_AVISO_FECHAS


def agregar_fecha_lanzamiento(listado: pd.DataFrame, cliente=None) -> pd.DataFrame:
    """Agrega la columna "fecha_lanzamiento" al listado de canciones,
    resuelta vía Spotify a partir de "isrc" (ver
    spotify_release_dates.resolver_fechas_lanzamiento).

    A propósito NO deja que un problema con la API de Spotify (sin
    credenciales configuradas, sin internet, rate limit agotado, ISRC no
    encontrado, etc.) tumbe la generación del reporte completo: si algo
    falla, la columna queda en blanco para esta corrida y se vuelve a
    intentar la próxima vez (la caché ya resuelta no se pierde).
    """
    global _ULTIMO_AVISO_FECHAS
    _ULTIMO_AVISO_FECHAS = None

    listado = listado.copy()
    if listado.empty or "isrc" not in listado.columns:
        listado["fecha_lanzamiento"] = None
        return listado
    try:
        listado["fecha_lanzamiento"] = spotify_release_dates.resolver_fechas_lanzamiento(
            listado["isrc"], cliente=cliente
        )
    except Exception as e:
        _ULTIMO_AVISO_FECHAS = (
            f"No se pudieron resolver las fechas de lanzamiento vía Spotify ({e}). "
            "El reporte se generó igual, con esa columna vacía; se vuelve a intentar "
            "en la próxima corrida."
        )
        print(f"Aviso: {_ULTIMO_AVISO_FECHAS}")
        listado["fecha_lanzamiento"] = None
    return listado


def _color_semaforo(banda: int, conteo):
    """Semáforo de participación de Universal para una celda conteo_universal
    de la serie histórica (banda/país/semana): el objetivo es que Universal
    tenga config.PCT_OBJETIVO_UNIVERSAL (30%) de los tracks de esa banda.

    - Rojo: por debajo del objetivo (incluye 0 -- el usuario dio el ejemplo
      con 1-2 canciones para banda=10, no mencionó el caso de 0, se trata
      igual de rojo por estar aún más lejos del objetivo. Avisar si no es
      lo esperado).
    - Amarillo: exactamente en el objetivo (3 de 10, 9 de 30, 15 de 50, 30
      de 100, 60 de 200).
    - Verde: por encima del objetivo.

    None (sin color) si no hay dato (celda vacía / NaN).
    """
    if conteo is None or (isinstance(conteo, float) and pd.isna(conteo)):
        return None
    objetivo = round(banda * config.PCT_OBJETIVO_UNIVERSAL)
    if conteo < objetivo:
        return _RELLENO_ROJO
    if conteo == objetivo:
        return _RELLENO_AMARILLO
    return _RELLENO_VERDE


def _escribir_bloques_pais(ws, fila_header_pais: int, fila_header_banda: int, aplicar_anchos: bool):
    """Escribe un juego de encabezados de país/banda (celda de país
    combinada arriba, bandas 10/30/50/100/200 debajo, columna angosta de
    separación entre países) empezando en _COL_PRIMER_PAIS -- lo usan tanto
    la serie histórica (filas 5/6) como el listado de canciones (sus propias
    filas de encabezado), para que ambos bloques queden alineados en las
    mismas columnas.

    `aplicar_anchos=False` evita volver a fijar el ancho de columnas ya
    fijado la primera vez (los mismos países/bandas caen siempre en las
    mismas columnas, así que no hace falta repetirlo).

    Devuelve (columna_inicio_por_pais, columna_siguiente_libre) --
    columna_siguiente_libre es la primera columna después del último país
    (después de su separadora), útil para ubicar columnas extra a la
    derecha del último bloque.
    """
    negrita_centrada = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    columna_inicio_por_pais: dict[str, int] = {}
    columna_actual = _COL_PRIMER_PAIS
    for pais in config.ORDEN_PAISES_CHART:
        columna_inicio_por_pais[pais] = columna_actual
        columna_fin_bloque = columna_actual + len(config.BANDAS_CHART) - 1

        nombre_visible = config.NOMBRE_PAIS_CHART[pais]
        celda_pais = ws.cell(row=fila_header_pais, column=columna_actual, value=nombre_visible)
        ws.merge_cells(
            start_row=fila_header_pais, start_column=columna_actual,
            end_row=fila_header_pais, end_column=columna_fin_bloque,
        )
        celda_pais.alignment = centrado
        celda_pais.font = negrita_centrada

        for i, banda in enumerate(config.BANDAS_CHART):
            celda_banda = ws.cell(row=fila_header_banda, column=columna_actual + i, value=banda)
            celda_banda.alignment = centrado
            celda_banda.font = negrita_centrada
            if aplicar_anchos:
                ws.column_dimensions[get_column_letter(columna_actual + i)].width = 4

        columna_separadora = columna_fin_bloque + 1
        if aplicar_anchos:
            ws.column_dimensions[get_column_letter(columna_separadora)].width = 1.5
        columna_actual = columna_separadora + 1

    return columna_inicio_por_pais, columna_actual


def _escribir_resumen_total(
    writer: pd.ExcelWriter, resumen: pd.DataFrame, df_semana: pd.DataFrame, cliente_spotify=None,
) -> None:
    """Escribe la pestaña "Resumen Total": arriba la serie histórica
    completa con el formato de la plantilla original (encabezados
    combinados, columnas angostas, freeze_panes), y debajo -- después de un
    salto de filas -- el listado de canciones de la semana actual (ver
    construir_listado_canciones). Crea la hoja directo con openpyxl
    (writer.book.create_sheet), en vez de con pandas .to_excel, porque acá
    se necesita control celda por celda que .to_excel(...) no ofrece. (Un
    ExcelWriter nuevo de pandas NO deja una hoja en blanco lista para usar
    -- writer.book.active es None hasta que se crea una hoja.)
    """
    ws = writer.book.create_sheet(config.CHART_SHEET_RESUMEN)

    negrita_centrada = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # "SEMANA / TOP" -- combinada verticalmente en las dos filas de encabezado
    celda_semana = ws.cell(row=_FILA_HEADER_PAIS, column=_COL_SEMANA, value="SEMANA \n/ TOP")
    ws.merge_cells(
        start_row=_FILA_HEADER_PAIS, start_column=_COL_SEMANA,
        end_row=_FILA_HEADER_BANDA, end_column=_COL_SEMANA,
    )
    celda_semana.alignment = centrado
    celda_semana.font = negrita_centrada

    # Columna A ("año" en la serie histórica, y "Artist/Título" en el
    # listado de canciones de más abajo) mucho más ancha de lo que necesita
    # un año solo, porque tiene que alcanzar para los títulos largos: desde
    # que la fecha de lanzamiento se movió a la columna B (a pedido del
    # usuario, ver _escribir_listado_canciones), los títulos ya no pueden
    # desbordar visualmente hacia B como antes.
    ws.column_dimensions[get_column_letter(_COL_ANIO)].width = 32
    ws.column_dimensions[get_column_letter(_COL_MES)].width = 12
    ws.column_dimensions[get_column_letter(_COL_SEMANA)].width = 10
    ws.column_dimensions[get_column_letter(_COL_SEPARADOR_INICIAL)].width = 2

    columna_inicio_por_pais, columna_siguiente_libre = _escribir_bloques_pais(
        ws, _FILA_HEADER_PAIS, _FILA_HEADER_BANDA, aplicar_anchos=True
    )

    # Filas de datos de la serie histórica: el año solo se escribe la
    # primera vez que aparece (igual que la plantilla original -- no está
    # combinado, solo se deja en blanco en las filas siguientes del mismo
    # año).
    anio_anterior = None
    for i, fila in enumerate(resumen.itertuples(index=False)):
        r = _FILA_PRIMER_DATO + i
        anio = int(fila.anio)
        if anio != anio_anterior:
            ws.cell(row=r, column=_COL_ANIO, value=anio)
            anio_anterior = anio
        ws.cell(row=r, column=_COL_MES, value=fila.mes)
        ws.cell(row=r, column=_COL_SEMANA, value=int(fila.semana))

        for pais in config.ORDEN_PAISES_CHART:
            col_inicio = columna_inicio_por_pais[pais]
            for i_banda, banda in enumerate(config.BANDAS_CHART):
                nombre_columna = f"{pais}_top{banda}"
                valor = getattr(fila, nombre_columna, None)
                celda = ws.cell(row=r, column=col_inicio + i_banda, value=valor)
                relleno = _color_semaforo(banda, valor)
                if relleno is not None:
                    celda.fill = relleno

    ws.freeze_panes = f"{get_column_letter(_COL_PRIMER_PAIS)}{_FILA_PRIMER_DATO}"

    ultima_fila_historica = _FILA_PRIMER_DATO + max(len(resumen) - 1, 0)
    fila_listado = ultima_fila_historica + _FILAS_ANTES_DE_LISTADO + 1
    _escribir_listado_canciones(
        ws, df_semana, fila_listado, columna_inicio_por_pais, columna_siguiente_libre,
        cliente_spotify=cliente_spotify,
    )


def _escribir_listado_canciones(
    ws, df_semana: pd.DataFrame, fila_inicio: int,
    columna_inicio_por_pais: dict, columna_siguiente_libre: int,
    cliente_spotify=None,
) -> None:
    """Escribe el bloque "listado de canciones" de la semana actual, debajo
    de la serie histórica en la misma pestaña "Resumen Total" (ver
    construir_listado_canciones para el detalle de qué contiene cada fila).
    No lleva histórico: se reescribe por completo cada vez que se genera el
    reporte, con la semana que se acaba de cargar.

    `cliente_spotify` es opcional -- se pasa tal cual a
    agregar_fecha_lanzamiento (ver ahí): None usa la API real de Spotify
    (con degradación segura si no hay credenciales o falla), o se puede
    inyectar un doble de prueba.
    """
    if df_semana.empty:
        return
    listado = construir_listado_canciones(df_semana)
    if listado.empty:
        return
    listado = agregar_fecha_lanzamiento(listado, cliente=cliente_spotify)

    negrita = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fecha = pd.Timestamp(df_semana["chart_date"].iloc[0])
    texto_titulo = f"Week Ending - {fecha.day:02d} {config.MESES_ES_ABREV[fecha.month]}, {fecha.year}"
    celda_titulo = ws.cell(row=fila_inicio, column=_COL_ANIO, value=texto_titulo)
    celda_titulo.font = negrita

    fila_header_pais = fila_inicio + 1
    fila_header_banda = fila_header_pais + 1
    fila_primer_dato = fila_header_banda + 1

    celda_cancion = ws.cell(row=fila_header_pais, column=_COL_ANIO, value="Artist/Título")
    ws.merge_cells(
        start_row=fila_header_pais, start_column=_COL_ANIO,
        end_row=fila_header_banda, end_column=_COL_ANIO,
    )
    celda_cancion.alignment = centrado
    celda_cancion.font = negrita

    # Fecha de lanzamiento (vía Spotify, ver agregar_fecha_lanzamiento) al
    # lado del nombre de la canción, antes de "Región" -- así lo pidió el
    # usuario mostrando la plantilla real (antes estaba al final, después de
    # "Suma Posiciones"). Va en _COL_MES (columna B), que en las filas del
    # listado está libre: así NO se corren los bloques de país, que tienen
    # que seguir alineados con los de la serie histórica de arriba (ambos
    # usan las mismas columnas, ver _escribir_bloques_pais).
    celda_fecha_lanz = ws.cell(row=fila_header_pais, column=_COL_MES, value="Fecha Lzto")
    ws.merge_cells(
        start_row=fila_header_pais, start_column=_COL_MES,
        end_row=fila_header_banda, end_column=_COL_MES,
    )
    celda_fecha_lanz.alignment = centrado
    celda_fecha_lanz.font = negrita

    celda_region = ws.cell(row=fila_header_pais, column=_COL_SEMANA, value="Región")
    ws.merge_cells(
        start_row=fila_header_pais, start_column=_COL_SEMANA,
        end_row=fila_header_banda, end_column=_COL_SEMANA,
    )
    celda_region.alignment = centrado
    celda_region.font = negrita

    # Mismos encabezados de país/banda que la serie histórica, en las mismas
    # columnas (no vuelve a fijar anchos -- ya quedaron fijos arriba).
    _escribir_bloques_pais(ws, fila_header_pais, fila_header_banda, aplicar_anchos=False)

    # Las dos columnas sin nombre que traía la plantilla original, al final
    # de los bloques de país: cuántos países tiene esa canción, y la suma de
    # sus posiciones en todos ellos.
    col_paises = columna_siguiente_libre
    col_suma = columna_siguiente_libre + 1

    celda_paises = ws.cell(row=fila_header_pais, column=col_paises, value="N° Países")
    ws.merge_cells(
        start_row=fila_header_pais, start_column=col_paises,
        end_row=fila_header_banda, end_column=col_paises,
    )
    celda_paises.alignment = centrado
    celda_paises.font = negrita
    ws.column_dimensions[get_column_letter(col_paises)].width = 9

    celda_suma = ws.cell(row=fila_header_pais, column=col_suma, value="Suma Posiciones")
    ws.merge_cells(
        start_row=fila_header_pais, start_column=col_suma,
        end_row=fila_header_banda, end_column=col_suma,
    )
    celda_suma.alignment = centrado
    celda_suma.font = negrita
    ws.column_dimensions[get_column_letter(col_suma)].width = 9

    for i, fila in enumerate(listado.itertuples(index=False)):
        r = fila_primer_dato + i
        ws.cell(row=r, column=_COL_ANIO, value=fila.cancion)
        fecha_lanzamiento = getattr(fila, "fecha_lanzamiento", None)
        if pd.notna(fecha_lanzamiento):
            # Texto, no fecha de Excel a propósito -- Spotify a veces solo
            # trae año o año-mes (release_date_precision), forzar un
            # number_format de fecha rompería esos casos parciales.
            ws.cell(row=r, column=_COL_MES, value=str(fecha_lanzamiento))
        ws.cell(row=r, column=_COL_SEMANA, value=fila.region)

        for pais in config.ORDEN_PAISES_CHART:
            col_inicio = columna_inicio_por_pais[pais]
            for i_banda, banda in enumerate(config.BANDAS_CHART):
                nombre_columna = f"{pais}_top{banda}"
                valor = getattr(fila, nombre_columna, None)
                if pd.notna(valor):
                    ws.cell(row=r, column=col_inicio + i_banda, value=int(valor))

        ws.cell(row=r, column=col_paises, value=int(fila.paises_presente))
        ws.cell(row=r, column=col_suma, value=int(fila.suma_posiciones))


# Layout de "Detalle Tracks": posición (1-200) fija a la izquierda y un
# país por columna, con la canción que ocupa esa posición en ese país esa
# semana ("Título / Artista", igual que el listado de canciones de "Resumen
# Total"). Fila de título (semana/fecha) y fila de encabezado de país fijas
# arriba con freeze_panes -- igual idea que "Resumen Total", pero sin las
# bandas 10/30/50/100/200 (acá cada país es una sola columna, no un bloque)
# y sin color todavía (el usuario no tiene definido qué deberían significar
# los colores de esta pestaña, se deja para más adelante).
_DETALLE_COL_POSICION = 1
_DETALLE_COL_PRIMER_PAIS = 2
_DETALLE_FILA_TITULO = 1
_DETALLE_FILA_HEADER_PAIS = 2
_DETALLE_FILA_PRIMER_DATO = 3


def construir_detalle_tracks(df_semana: pd.DataFrame) -> pd.DataFrame:
    """Detalle track por track de la semana que se acaba de cargar: una fila
    por posición (1-200), una columna por país (config.ORDEN_PAISES_CHART),
    con el texto "Título / Artista" de la canción que ocupa esa posición en
    ese país. Sin histórico — ver advertencia del módulo.

    Antes esta función devolvía una tabla larga (una fila por país+posición,
    con columnas country_code/chart_date/position/artist/song_name/
    stream_count/label_group/label_name); se cambió a esta tabla ancha para
    que la pestaña se vea como la plantilla original -- OJO, con el cambio
    se dejan de mostrar stream_count/label_group/label_name (no se pidieron
    para este formato); avisar si hacen falta en otro lado.
    """
    if df_semana.empty:
        return pd.DataFrame(columns=["position"])

    df = df_semana.copy()
    df["cancion"] = (
        df["song_name"].astype(str).str.strip() + " / " + df["artist"].astype(str).str.strip()
    )
    tabla = df.pivot_table(index="position", columns="country_code", values="cancion", aggfunc="first")
    columnas_ordenadas = [pais for pais in config.ORDEN_PAISES_CHART if pais in tabla.columns]
    tabla = tabla[columnas_ordenadas]
    tabla = tabla.reset_index().sort_values("position").reset_index(drop=True)
    return tabla


def _escribir_detalle_tracks(ws, df_semana: pd.DataFrame, tabla: pd.DataFrame) -> None:
    """Escribe la pestaña "Detalle Tracks" con el layout de arriba
    (_DETALLE_COL_*/_DETALLE_FILA_*) -- se crea la hoja aparte con openpyxl
    en vez de con `.to_excel(...)` para poder fijar (freeze_panes) la
    columna de posición y la fila de encabezado, igual que en "Resumen
    Total".
    """
    if tabla.empty:
        return

    negrita = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fecha = pd.Timestamp(df_semana["chart_date"].iloc[0])
    texto_titulo = f"Week Ending - {fecha.day:02d} {config.MESES_ES_ABREV[fecha.month]}, {fecha.year}"
    celda_titulo = ws.cell(row=_DETALLE_FILA_TITULO, column=_DETALLE_COL_POSICION, value=texto_titulo)
    celda_titulo.font = negrita

    celda_posicion = ws.cell(row=_DETALLE_FILA_HEADER_PAIS, column=_DETALLE_COL_POSICION, value="Position")
    celda_posicion.font = negrita
    celda_posicion.alignment = centrado
    ws.column_dimensions[get_column_letter(_DETALLE_COL_POSICION)].width = 10

    columnas_pais = [c for c in tabla.columns if c != "position"]
    for i, pais in enumerate(columnas_pais):
        col = _DETALLE_COL_PRIMER_PAIS + i
        celda = ws.cell(row=_DETALLE_FILA_HEADER_PAIS, column=col, value=pais)
        celda.font = negrita
        celda.alignment = centrado
        ws.column_dimensions[get_column_letter(col)].width = 24

    for i, fila in enumerate(tabla.itertuples(index=False)):
        r = _DETALLE_FILA_PRIMER_DATO + i
        ws.cell(row=r, column=_DETALLE_COL_POSICION, value=int(fila.position))
        for j, pais in enumerate(columnas_pais):
            valor = getattr(fila, pais, None)
            if pd.notna(valor):
                ws.cell(row=r, column=_DETALLE_COL_PRIMER_PAIS + j, value=valor)

    ws.freeze_panes = f"{get_column_letter(_DETALLE_COL_PRIMER_PAIS)}{_DETALLE_FILA_PRIMER_DATO}"


def generar_reporte(
    df_semana: pd.DataFrame, output_path: Path, guardar_en_historico: bool = True,
    cliente_spotify=None,
) -> Path:
    """Genera el reporte de Chart Semanal a partir del DataFrame de la
    semana nueva (el que devuelve load_data.load_source). Guarda esa semana
    en el histórico (a menos que ya se haya guardado antes, con
    guardar_en_historico=False) -- tanto el conteo agregado
    (`chart_band_weekly`, para "Resumen Total") como el detalle track por
    track (`chart_track_weekly`, para no perder "Detalle Tracks"/el listado
    de canciones de esta semana aunque el reporte solo muestre la semana
    actual) -- y arma "Resumen Total" con TODO el histórico acumulado más el
    listado de canciones de la semana actual (incluida su columna "Fecha de
    Lanzamiento", resuelta vía Spotify -- ver agregar_fecha_lanzamiento), y
    "Detalle Tracks" solo con la semana actual (posición fija x país por
    columna, ver construir_detalle_tracks / _escribir_detalle_tracks).

    `cliente_spotify` es opcional: por defecto (None) se usa la API real de
    Spotify (credenciales desde el entorno, con degradación segura si
    faltan o algo falla -- ver agregar_fecha_lanzamiento). Se puede pasar
    un doble de prueba para no depender de la red/credenciales reales.
    """
    if guardar_en_historico:
        history.append_semana_chart(df_semana)
        history.append_semana_tracks(df_semana)

    resumen = construir_resumen_total()
    detalle = construir_detalle_tracks(df_semana)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _escribir_resumen_total(writer, resumen, df_semana, cliente_spotify=cliente_spotify)
        ws_detalle = writer.book.create_sheet(config.CHART_SHEET_DETALLE)
        _escribir_detalle_tracks(ws_detalle, df_semana, detalle)

    return output_path
