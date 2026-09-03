"""Pruebas de src/chart_semanal.py — usan una base SQLite temporal (nunca
tocan data/history/universal_data.db de verdad).

Corre con: pytest tests/test_chart_semanal.py -v
"""
import openpyxl
import pandas as pd
import pytest

from src import chart_semanal, config, history


@pytest.fixture(autouse=True)
def db_temporal(tmp_path, monkeypatch):
    """Redirige history.DB_PATH a un archivo temporal por cada test, para no
    tocar nunca la base real del proyecto."""
    monkeypatch.setattr(history, "DB_PATH", tmp_path / "test_universal_data.db")


def _csv_vacio(tmp_path, nombre, columnas):
    path = tmp_path / nombre
    pd.DataFrame(columns=columnas).to_csv(path, index=False)
    return path


def test_construir_resumen_total_pivotea_columnas_pais_banda(tmp_path):
    chart_csv = tmp_path / "seed_chart.csv"
    pd.DataFrame([
        {"anio": 2026, "semana": 1, "mes": "ENERO", "country_code": "CO", "banda": 10, "conteo_universal": 1},
        {"anio": 2026, "semana": 1, "mes": "ENERO", "country_code": "CO", "banda": 30, "conteo_universal": 5},
        {"anio": 2026, "semana": 1, "mes": "ENERO", "country_code": "PE", "banda": 10, "conteo_universal": 2},
    ]).to_csv(chart_csv, index=False)
    ms_csv = _csv_vacio(tmp_path, "seed_ms.csv",
                        ["anio", "semana", "country_code", "label_group", "streams_top200", "chart_date"])
    history.seed_historico(chart_csv, ms_csv)

    resumen = chart_semanal.construir_resumen_total()

    assert list(resumen.columns[:3]) == ["anio", "semana", "mes"]
    assert "CO_top10" in resumen.columns and "PE_top10" in resumen.columns
    fila = resumen.iloc[0]
    assert fila["CO_top10"] == 1
    assert fila["CO_top30"] == 5
    assert fila["PE_top10"] == 2


def test_construir_resumen_total_vacio_no_falla(tmp_path):
    resumen = chart_semanal.construir_resumen_total()
    assert resumen.empty


def test_append_semana_chart_guarda_mes_en_espanol_no_en_ingles(tmp_path):
    # Regresión del bug real que encontramos: fecha.strftime("%B") depende
    # del locale del sistema y puede devolver "JUNE" en vez de "JUNIO",
    # rompiendo la continuidad con el histórico sembrado (que siempre está
    # en español). Verificado contra Reporte_Chart_Top Semanal Spotify Latam
    # a Sem 24 de 2026.xlsm, donde todas las filas usan mes en español.
    chart_csv = _csv_vacio(tmp_path, "seed_chart.csv",
                           ["anio", "semana", "mes", "country_code", "banda", "conteo_universal"])
    ms_csv = _csv_vacio(tmp_path, "seed_ms.csv",
                        ["anio", "semana", "country_code", "label_group", "streams_top200", "chart_date"])
    history.seed_historico(chart_csv, ms_csv)

    df_semana = pd.DataFrame({
        "country_code": ["CO"],
        "label_group": ["Universal"],
        "position": [1],
        "stream_count": [1_000_000],
        "chart_date": pd.to_datetime(["2026-06-18"]),
    })
    history.append_semana_chart(df_semana)

    chart_df = history.cargar_chart_band_weekly()
    assert chart_df["mes"].iloc[0] == "JUNIO"


def test_construir_detalle_tracks_ordena_por_pais_y_posicion(tmp_path):
    df_semana = pd.DataFrame({
        "country_code": ["PE", "CO", "CO"],
        "chart_date": pd.to_datetime(["2026-06-18"] * 3),
        "position": [1, 2, 1],
        "artist": ["a", "b", "c"],
        "song_name": ["x", "y", "z"],
        "stream_count": [100, 200, 300],
        "label_group": ["Universal", "Sony", "Universal"],
        "label_name": ["UMG", "Sony Music", "UMG"],
    })
    detalle = chart_semanal.construir_detalle_tracks(df_semana)
    assert list(detalle["country_code"]) == ["CO", "CO", "PE"]
    assert list(detalle["position"]) == [1, 2, 1]


def test_conteo_universal_coincide_con_datos_reales_de_colombia_semana_24(tmp_path):
    # Validación Paso 6: estos son los 10 tracks REALES del Top 10 de Colombia
    # en la semana del 2026-06-11 (BASE Informe Sportify charts Semana 24),
    # tal cual venían en la fuente cruda (position, label_group). El conteo
    # "Universal" para TOP 10 en el reporte oficial (Reporte_Chart_Top Semanal
    # Spotify Latam a Sem 24 de 2026.xlsm, fila de la semana 24) es 1 -- y la
    # validación completa (17 países x 5 bandas = 85 valores, hecha aparte
    # contra BASE Informe Sportify charts Semana 24 y el reporte oficial)
    # confirmó que este método coincide exactamente en los 85 casos. Este
    # test deja un caso real fijo como regresión rápida.
    chart_csv = _csv_vacio(tmp_path, "seed_chart.csv",
                           ["anio", "semana", "mes", "country_code", "banda", "conteo_universal"])
    ms_csv = _csv_vacio(tmp_path, "seed_ms.csv",
                        ["anio", "semana", "country_code", "label_group", "streams_top200", "chart_date"])
    history.seed_historico(chart_csv, ms_csv)

    top10_reales_co = [
        (1, "Warner"), (2, "Warner"), (3, "INgrooves"), (4, "INgrooves"),
        (5, "Warner"), (6, "Warner"), (7, "Warner"), (8, "Sony"),
        (9, "Universal"), (10, "Warner"),
    ]
    df_semana = pd.DataFrame({
        "country_code": ["CO"] * 10,
        "label_group": [lbl for _, lbl in top10_reales_co],
        "position": [pos for pos, _ in top10_reales_co],
        "stream_count": [1_000_000] * 10,
        "chart_date": pd.to_datetime(["2026-06-11"] * 10),
    })
    history.append_semana_chart(df_semana)

    resumen = chart_semanal.construir_resumen_total()
    assert resumen.iloc[0]["CO_top10"] == 1  # valor oficial real verificado


def test_generar_reporte_escribe_las_dos_pestanas(tmp_path):
    chart_csv = _csv_vacio(tmp_path, "seed_chart.csv",
                           ["anio", "semana", "mes", "country_code", "banda", "conteo_universal"])
    ms_csv = _csv_vacio(tmp_path, "seed_ms.csv",
                        ["anio", "semana", "country_code", "label_group", "streams_top200", "chart_date"])
    history.seed_historico(chart_csv, ms_csv)

    df_semana = pd.DataFrame({
        "country_code": ["CO"],
        "chart_date": pd.to_datetime(["2026-06-18"]),
        "position": [1],
        "artist": ["a"],
        "song_name": ["x"],
        "stream_count": [1_000_000],
        "label_group": ["Universal"],
        "label_name": ["UMG"],
    })
    salida = tmp_path / "reporte.xlsx"
    chart_semanal.generar_reporte(df_semana, salida)

    assert salida.exists()
    wb = openpyxl.load_workbook(salida)
    assert wb.sheetnames == [config.CHART_SHEET_RESUMEN, config.CHART_SHEET_DETALLE]
    detalle = pd.read_excel(salida, sheet_name=config.CHART_SHEET_DETALLE)
    assert len(detalle) == 1


def test_resumen_total_replica_el_formato_de_la_plantilla_original(tmp_path):
    # Paso 7 (formato): freeze panes en las columnas/filas fijas, país
    # combinado en la fila 5 con el orden y nombre real de la plantilla,
    # bandas (10/30/50/100/200) en la fila 6, columna angosta de separación
    # entre cada país -- verificado 1:1 contra Reporte_Chart_Top Semanal
    # Spotify Latam a Sem 24 de 2026.xlsm (misma fila 5/6, freeze_panes E7,
    # 18 celdas combinadas).
    chart_csv = _csv_vacio(tmp_path, "seed_chart.csv",
                           ["anio", "semana", "mes", "country_code", "banda", "conteo_universal"])
    ms_csv = _csv_vacio(tmp_path, "seed_ms.csv",
                        ["anio", "semana", "country_code", "label_group", "streams_top200", "chart_date"])
    history.seed_historico(chart_csv, ms_csv)

    df_semana = pd.DataFrame({
        "country_code": ["CO", "PE"],
        "chart_date": pd.to_datetime(["2026-06-18"] * 2),
        "position": [1, 1],
        "artist": ["a", "b"],
        "song_name": ["x", "y"],
        "stream_count": [1_000_000, 1_000_000],
        "label_group": ["Universal", "Universal"],
        "label_name": ["UMG", "UMG"],
    })
    salida = tmp_path / "reporte.xlsx"
    chart_semanal.generar_reporte(df_semana, salida)

    wb = openpyxl.load_workbook(salida)
    ws = wb[config.CHART_SHEET_RESUMEN]

    assert ws.freeze_panes == "E7"
    assert ws.cell(row=5, column=3).value == "SEMANA \n/ TOP"
    assert ws.cell(row=5, column=5).value == "COLOMBIA"  # primer bloque de país (col E)
    assert ws.cell(row=5, column=11).value == "PERU"  # segundo bloque (col K) -- 1 col de separación (J)
    assert [ws.cell(row=6, column=c).value for c in range(5, 10)] == [10, 30, 50, 100, 200]
    assert ws.cell(row=6, column=10).value is None  # columna J: separadora, vacía
    assert ws.column_dimensions["J"].width < 3  # separadora, angosta
    # merges de la fila 5 nada más (la pestaña tiene más celdas combinadas
    # abajo, del listado de canciones -- ver test aparte).
    merges_fila_5 = [r for r in ws.merged_cells.ranges if r.min_row == 5]
    assert len(merges_fila_5) == 18  # 1 (SEMANA/TOP) + 17 países


def test_construir_listado_canciones_arma_una_fila_por_cancion_con_posicion_por_pais_y_totales(tmp_path):
    # "Coleccionando Heridas" queda en el Top 10 de Colombia (posición 5) y
    # en el Top 30 de Perú (posición 20) -- dos países, suma de posiciones
    # 25. "Golden" solo aparece en Perú, fuera del Top 200 (no debe listarse
    # con ninguna banda).
    df_semana = pd.DataFrame({
        "country_code": ["CO", "PE", "PE"],
        "chart_date": pd.to_datetime(["2026-06-18"] * 3),
        "position": [5, 20, 250],
        "artist": ["KAROL G", "KAROL G", "HUNTR/X"],
        "song_name": ["Coleccionando Heridas", "Coleccionando Heridas", "Golden"],
        "stream_count": [1_000_000] * 3,
        "label_group": ["Universal"] * 3,
        "label_name": ["UMG"] * 3,
        "region": ["Latin", "Latin", "Anglo"],
    })

    listado = chart_semanal.construir_listado_canciones(df_semana)

    assert len(listado) == 1  # "Golden" quedó fuera (posición 250, fuera del Top 200)
    fila = listado.iloc[0]
    assert fila["cancion"] == "Coleccionando Heridas / KAROL G"
    assert fila["region"] == "Latin"
    assert fila["paises_presente"] == 2
    assert fila["suma_posiciones"] == 25
    assert fila["CO_top10"] == 5  # posición 5 -> banda 10
    assert fila["PE_top30"] == 20  # posición 20 -> banda 30
    assert pd.isna(fila.get("PE_top10"))  # no le corresponde esa banda en Perú


def test_listado_canciones_vacio_no_falla(tmp_path):
    assert chart_semanal.construir_listado_canciones(pd.DataFrame()).empty


def test_construir_listado_canciones_se_recorta_a_las_mejores_200(tmp_path, monkeypatch):
    # Con una semana real puede haber 1000+ canciones distintas -- el
    # usuario pidió dejar solo las mejores 200 (mismo orden: más países y
    # mejor suma de posiciones primero). Se prueba con un límite más chico
    # (5) para no armar 250 filas de datos a mano.
    monkeypatch.setattr(config, "TOP_N_LISTADO_CANCIONES", 5)
    n_canciones = 8
    df_semana = pd.DataFrame({
        "country_code": ["CO"] * n_canciones,
        "chart_date": pd.to_datetime(["2026-06-18"] * n_canciones),
        "position": list(range(1, n_canciones + 1)),
        "artist": [f"artista{i}" for i in range(n_canciones)],
        "song_name": [f"cancion{i}" for i in range(n_canciones)],
        "stream_count": [1_000_000] * n_canciones,
        "label_group": ["Universal"] * n_canciones,
        "label_name": ["UMG"] * n_canciones,
        "region": ["Latin"] * n_canciones,
    })

    listado = chart_semanal.construir_listado_canciones(df_semana)

    assert len(listado) == 5
    # las 5 mejores son las de menor suma de posiciones (todas con 1 solo
    # país, así que desempata por suma_posiciones ascendente)
    assert list(listado["suma_posiciones"]) == [1, 2, 3, 4, 5]


def test_resumen_total_incluye_el_listado_de_canciones_debajo_de_la_serie_historica(tmp_path):
    chart_csv = _csv_vacio(tmp_path, "seed_chart.csv",
                           ["anio", "semana", "mes", "country_code", "banda", "conteo_universal"])
    ms_csv = _csv_vacio(tmp_path, "seed_ms.csv",
                        ["anio", "semana", "country_code", "label_group", "streams_top200", "chart_date"])
    history.seed_historico(chart_csv, ms_csv)

    df_semana = pd.DataFrame({
        "country_code": ["CO", "PE"],
        "chart_date": pd.to_datetime(["2026-06-18"] * 2),
        "position": [1, 40],
        "artist": ["a", "b"],
        "song_name": ["x", "y"],
        "stream_count": [1_000_000, 1_000_000],
        "label_group": ["Universal", "Universal"],
        "label_name": ["UMG", "UMG"],
        "region": ["Latin", "Anglo"],
    })
    salida = tmp_path / "reporte.xlsx"
    chart_semanal.generar_reporte(df_semana, salida)

    wb = openpyxl.load_workbook(salida)
    ws = wb[config.CHART_SHEET_RESUMEN]

    # Con una sola semana histórica, la serie ocupa la fila 7; el listado
    # empieza 2 filas en blanco después (filas 8-9), título en la fila 10,
    # encabezados en 11/12, primera canción en la 13.
    assert ws.cell(row=10, column=1).value == "Week Ending - 18 jun, 2026"
    assert ws.cell(row=11, column=1).value == "Artist/Título"
    assert ws.cell(row=11, column=3).value == "Región"
    assert ws.cell(row=11, column=5).value == "COLOMBIA"
    assert [ws.cell(row=12, column=c).value for c in range(5, 10)] == [10, 30, 50, 100, 200]
    # columnas finales sin nombre en la plantilla original -> las nombramos.
    # Columna 107 = después del bloque 17 (países) x 6 (5 bandas + 1
    # separadora) empezando en la columna 5.
    col_paises, col_suma = 107, 108
    assert ws.cell(row=11, column=col_paises).value == "N° Países"
    assert ws.cell(row=11, column=col_suma).value == "Suma Posiciones"

    assert ws.cell(row=13, column=1).value == "x / a"
    assert ws.cell(row=13, column=3).value == "Latin"
    assert ws.cell(row=13, column=5).value == 1  # CO, banda 10
    assert ws.cell(row=13, column=col_paises).value == 1
    assert ws.cell(row=13, column=col_suma).value == 1
